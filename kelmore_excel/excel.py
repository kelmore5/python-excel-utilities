from __future__ import annotations

import csv
import os
import xml.etree.ElementTree
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Type, Tuple

import xlrd
from kelmore_files import FileTools as Files
from kelmore_strings import StringTools as Strings
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import Worksheet
from xlrd import Book
from xlrd.sheet import Sheet

Matrix = List[List[str]]
ExcelRows = Dict[str, Matrix]
ExcelJSON = Dict[str, List[Dict[str, str]]]


class ExcelPath:

    @staticmethod
    def files(directory: str,
              extension: str = '.xlsx',
              include_file_path: bool = False) -> List[str]:
        if not Files.check.is_directory(directory):
            raise IOError('The given file path must be a directory')

        return Files.directories.files(directory,
                                       include_full_path=include_file_path,
                                       extension=extension)

    @staticmethod
    def is_excel_file(full_path: str) -> bool:
        return Files.check.is_file(full_path) and ExcelPath.is_valid(full_path)

    @staticmethod
    def is_valid(full_path: str) -> bool:
        return full_path.endswith('.xlsx')


class ExcelSpecial:

    @staticmethod
    def excel_date_to_datetime(excel_date, date_mode):
        """ prec: excel_date is an excel datetime, date_mode is the date_mode
                    (0 for 1900-based, 1 for 1904 based)
            postc: returns datetime as a string
        """
        # date_mode: 0 for 1900-based, 1 for 1904-based
        return datetime(1899, 12, 30) + timedelta(days=excel_date + 1462 * date_mode)

    @staticmethod
    def remove_illegal_characters(value: Optional[str]) -> str:
        """ prec: value is a string
            postc: returns the string without unicode characters above index 128"""
        if not value:
            return ''

        return Strings.convert.only_ascii(value)

    @staticmethod
    def remove_illegal_characters_from_row(row: List[str]) -> List[str]:
        """ prec: value is a string
            postc: returns the string without unicode characters above index 128"""
        return [ExcelSpecial.remove_illegal_characters(x) for x in row]

    @staticmethod
    def remove_illegal_characters_from_rows(rows: Matrix) -> Matrix:
        """ prec: value is a string
            postc: returns the string without unicode characters above index 128"""
        return [ExcelSpecial.remove_illegal_characters_from_row(x) for x in rows]


class ExcelTransform:

    @staticmethod
    def to_csv(excel_path: str,
               csv_path: Optional[str] = None,
               overwrite: bool = False) -> List[str]:
        """ prec: file is a valid excel file path
            postc: creates a csv file from the given excel file"""
        if not csv_path:
            split_path: Tuple[str, str] = os.path.splitext(excel_path)
            csv_path = f'{split_path[0]}.csv'

        if not overwrite and Files.check.exists(csv_path):
            raise IOError('The given CSV path already exists')

        output: List[str] = []

        sheets: ExcelRows = ExcelTransform.to_rows(excel_path)
        for idx, sheet in enumerate(sheets):
            split_csv: Tuple[str, str] = os.path.splitext(csv_path)
            csv_file_name: str = f'{split_csv[0]}_{idx}.csv'

            with Files.io_.open(csv_file_name, write=True) as csv_file:
                csv_writer = csv.writer(csv_file, delimiter=",", quoting=csv.QUOTE_ALL)
                rows: List[List[str]] = sheets[sheet]
                for row in rows:
                    csv_writer.writerow(row)

            output.append(csv_file_name)

        return output

    @staticmethod
    def to_json(full_path: str) -> ExcelJSON:
        sheets: ExcelRows = ExcelTransform.to_rows(full_path)
        output: ExcelJSON = {}

        for sheet in sheets:
            sheet_as_json: List[Dict[str, str]] = []
            rows: List[List[str]] = sheets[sheet]

            headers: List[str] = rows.pop(0)
            for row in rows:
                new_row: Dict[str, str] = {}

                while len(row) < len(headers):
                    row.append('')

                for idx, header in enumerate(headers):
                    new_row[header] = row[idx]

                sheet_as_json.append(new_row)

            output[sheet] = sheet_as_json

        return output

    @staticmethod
    def to_rows(full_path: str) -> ExcelRows:
        if not ExcelPath.is_valid(full_path):
            raise ValueError('The given Excel path was not an Excel file')

        if not Files.check.exists(full_path):
            raise ValueError('The given Excel path did not exist')

        workbook: Book = xlrd.open_workbook(full_path)
        sheets: List[Sheet] = workbook.sheets()

        output: ExcelRows = {}
        for sheet in sheets:
            new_sheet: Matrix = []

            for row_idx in range(sheet.nrows):
                new_sheet.append(sheet.row_values(row_idx))

            output[sheet.name] = new_sheet

        return output

    @staticmethod
    def xml_to_excel(xml_path: str,
                     excel_path: str,
                     overwrite: bool = False) -> None:
        """ prec: xml_path is an xml file, excel_path is the name of the
                    spreadsheet to be created
            postc: creates an excel file from the given xml file (tags are columns, text is rows)
        """
        if not Files.check.exists(xml_path):
            raise IOError('The given XML file does not exist')
        rows: List[List[str]] = [[]]

        root = xml.etree.ElementTree.parse(xml_path).getroot()
        for child in root[0]:
            rows[0].append(child.tag)

        for child in root:
            new_row = []
            for tag in child:
                new_row.append(str(tag.text))
            rows.append(new_row)

        ExcelTools.create(excel_path, rows, overwrite=overwrite)


class ExcelTools:
    path: Type[ExcelPath] = ExcelPath
    special: Type[ExcelSpecial] = ExcelSpecial
    transform: Type[ExcelTransform] = ExcelTransform

    @staticmethod
    def create(full_path: str,
               rows: List[List[str]],
               initial_sheet_name: str = None,
               overwrite: bool = False) -> Workbook:
        """ prec: name is the name of the output file for the merged spreadsheet,
                    rows is the list of all the rows ready to be saved to the spreadsheet
            postc: merges the new rows from the excel file into the row dictionary.
                    Removes duplicates
        """
        if not overwrite and Files.check.exists(full_path):
            raise IOError('The given file path already exists')

        workbook: Workbook = Workbook()

        if initial_sheet_name:
            workbook.active.title = initial_sheet_name

        if rows:
            sheet: Worksheet = workbook.active
            for row_idx, row in enumerate(rows):
                for col_idx, value in enumerate(row):
                    value = ExcelSpecial.remove_illegal_characters(value)
                    sheet.cell(column=col_idx + 1, row=row_idx + 1, value=value)

        workbook.save(filename=full_path)
        return workbook

    @staticmethod
    def open(full_path: str) -> Workbook:
        return load_workbook(full_path)


class ExcelSheetWrapper:
    wrapper: ExcelWrapper
    sheet: Worksheet

    def __init__(self, wrapper: ExcelWrapper, sheet: Worksheet):
        self.wrapper = wrapper
        self.sheet = sheet

    def append(self, rows: Matrix) -> None:
        rows = ExcelTools.special.remove_illegal_characters_from_rows(rows)

        self.sheet.append(rows)
        self.wrapper.save()

    def clear(self) -> None:
        sheet_name: str = self.sheet.title
        sheet_idx: int = self.wrapper.workbook.sheetnames.index(sheet_name)

        self.wrapper.workbook.remove(self.sheet)
        self.sheet = self.wrapper.workbook.create_sheet(sheet_name, sheet_idx)

    def rows(self) -> Matrix:
        return [x for x in self.sheet]

    def row_count(self) -> int:
        return self.sheet.max_row

    def overwrite(self, rows: Matrix, row_start: int = 0) -> None:
        sheet = self.sheet

        for row_idx, row in enumerate(rows):
            for col_idx, value in enumerate(row):
                value = ExcelSpecial.remove_illegal_characters(value)
                sheet.cell(column=col_idx + 1, row=row_idx + 1 + row_start, value=value)

        self.wrapper.save()


class ExcelWrapper:
    full_path: str
    workbook: Workbook

    def __init__(self, full_path: str):
        if not ExcelPath.is_valid(full_path):
            raise ValueError('The given Excel path was not an Excel file')

        self.full_path = full_path
        self.workbook = ExcelTools.open(full_path) if Files.check.exists(full_path) else Workbook()

    @staticmethod
    def create(full_path: str, initial_sheet_name: str = None, overwrite: bool = False):
        if not overwrite and Files.check.exists(full_path):
            raise IOError('The given file path already exists, try open instead')

        wrapper: ExcelWrapper = ExcelWrapper(full_path)
        wrapper.workbook = ExcelTools.create(full_path,
                                             [],
                                             initial_sheet_name=initial_sheet_name,
                                             overwrite=overwrite)

        return wrapper

    @staticmethod
    def open(full_path: str) -> ExcelWrapper:
        if not Files.check.exists(full_path):
            raise IOError('The given file path did not exist, try create instead')

        return ExcelWrapper(full_path)

    def new_sheet(self, sheet_name: str) -> ExcelSheetWrapper:
        return ExcelSheetWrapper(self, self.workbook.create_sheet(title=sheet_name))

    def sheet(self, sheet_name: Optional[str]) -> ExcelSheetWrapper:
        sheet: Worksheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        return ExcelSheetWrapper(self, sheet)

    def save(self):
        self.workbook.save(filename=self.full_path)
